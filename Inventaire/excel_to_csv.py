from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "inventaire_parc_informatique.xlsx"
SHEET_NAME = "Inventaire"
CSV_OUT = BASE_DIR / "inventaire_export.csv"


def _excel_password() -> Optional[str]:
    pwd_file = BASE_DIR / "mdp_excel.txt"
    if not pwd_file.exists():
        return None
    lines = pwd_file.read_text(encoding="utf-8").splitlines()
    pwd = next((l.strip() for l in reversed(lines) if l.strip()), "")
    return pwd or None


def _load_workbook_keep_vba():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable: {EXCEL_PATH}")

    try:
        return load_workbook(EXCEL_PATH, keep_vba=False)
    except Exception:
        pwd = _excel_password()
        if not pwd:
            raise
        return load_workbook(EXCEL_PATH, keep_vba=False, password=pwd)


def _safe(v: Any) -> str:
    return "" if v is None else str(v)


def main() -> None:
    wb = None
    try:
        wb = _load_workbook_keep_vba()
        ws = wb[SHEET_NAME]

        header = [
            "Code-barres",
            "Type",
            "Marque",
            "Modele",
            "N° Serie",
            "Etat",
            "Localisation",
            "Utilisateur",
            "Date ajout",
            "Commentaire",
        ]

        # Writer CSV standard pour eviter les problemes de virgules dans les champs
        with open(CSV_OUT, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=",", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(header)

            max_row = ws.max_row or 0
            for r in range(2, max_row + 1):
                code = ws.cell(row=r, column=1).value
                if not code:
                    continue

                writer.writerow(
                    [
                        str(code).strip().upper(),
                        _safe(ws.cell(row=r, column=2).value),
                        _safe(ws.cell(row=r, column=3).value),
                        _safe(ws.cell(row=r, column=4).value),
                        _safe(ws.cell(row=r, column=5).value),
                        _safe(ws.cell(row=r, column=6).value),
                        _safe(ws.cell(row=r, column=7).value),
                        _safe(ws.cell(row=r, column=8).value),
                        _safe(ws.cell(row=r, column=9).value),
                        _safe(ws.cell(row=r, column=10).value),
                    ]
                )

        print(f"OK - CSV genere: {CSV_OUT.name}")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()

