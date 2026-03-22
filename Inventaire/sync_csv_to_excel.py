from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "inventaire_parc_informatique.xlsx"
SHEET_NAME = "Inventaire"
HEADER_ROW = 1


def _excel_password() -> Optional[str]:
    pwd_file = BASE_DIR / "mdp_excel.txt"
    if not pwd_file.exists():
        return None
    lines = pwd_file.read_text(encoding="utf-8").splitlines()
    pwd = next((l.strip() for l in reversed(lines) if l.strip()), "")
    return pwd or None


def _load_workbook_keep_vba():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable: {EXCEL_PATH}")

    try:
        return load_workbook(EXCEL_PATH, keep_vba=False)
    except Exception:
        pwd = _excel_password()
        if not pwd:
            raise
        return load_workbook(EXCEL_PATH, keep_vba=False, password=pwd)


def _norm_code(v: Any) -> str:
    return "" if v is None else str(v).strip().upper()


def _safe(v: Any) -> str:
    return "" if v is None else str(v)


def _read_csv_rows(csv_path: Path) -> List[List[str]]:
    raw = csv_path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")

    sample = text[:4096]
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t"])
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: List[List[str]] = []
    for row in reader:
        if not row:
            continue
        rows.append([str(col).strip() for col in row])
    return rows


def _map_headers(headers: List[str]) -> Dict[str, int]:
    h = [str(x).strip().lower() for x in headers]

    def find(substrs: List[str]) -> Optional[int]:
        for i, col in enumerate(h):
            if any(s in col for s in substrs):
                return i
        return None

    i_code = find(["code", "barres"])
    i_type = find(["type"])
    if i_code is None or i_type is None:
        return {
            "code_barre": 0,
            "type": 1,
            "marque": 2,
            "modele": 3,
            "numero_serie": 4,
            "etat": 5,
            "localisation": 6,
            "utilisateur": 7,
            "date_ajout": 8,
            "commentaire": 9,
        }

    def or_minus1(idx: Optional[int]) -> int:
        return -1 if idx is None else idx

    return {
        "code_barre": i_code,
        "type": i_type,
        "marque": or_minus1(find(["marque"])),
        "modele": or_minus1(find(["modele"])),
        "numero_serie": or_minus1(find(["serie"])),
        "etat": or_minus1(find(["etat"])),
        "localisation": or_minus1(find(["localisation", "local"])),
        "utilisateur": or_minus1(find(["utilisateur", "user"])),
        "date_ajout": or_minus1(find(["date"])),
        "commentaire": or_minus1(find(["commentaire", "comment"])),
    }


def _extract_row(row: List[str], idx_map: Dict[str, int]) -> Dict[str, str]:
    def get(key: str) -> str:
        idx = idx_map.get(key, -1)
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx]).strip()

    return {
        "code_barre": _norm_code(get("code_barre")),
        "type": get("type"),
        "marque": get("marque"),
        "modele": get("modele"),
        "numero_serie": get("numero_serie"),
        "etat": get("etat"),
        "localisation": get("localisation"),
        "utilisateur": get("utilisateur"),
        "date_ajout": get("date_ajout"),
        "commentaire": get("commentaire"),
    }


def _parse_date_like_excel(s: str) -> str:
    if not str(s).strip():
        return datetime.now().strftime("%d/%m/%Y %H:%M")
    return str(s).strip()


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Sync CSV vers inventaire .xlsm")
    parser.add_argument(
        "csv_path",
        nargs="?",
        default=str(BASE_DIR / "inventaire_export.csv"),
        help="Chemin vers le CSV genere par la page web",
    )
    args = parser.parse_args()
    csv_path = Path(args.csv_path)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV introuvable: {csv_path}")

    rows = _read_csv_rows(csv_path)
    if not rows:
        raise ValueError("CSV vide.")

    first = " ".join([str(x).lower() for x in rows[0] if x]).strip()
    has_headers = ("code" in first and "bar" in first) or any(
        "code" in str(x).lower() for x in rows[0]
    )

    start_idx = 1 if has_headers else 0
    idx_map = _map_headers(rows[0]) if has_headers else {
        "code_barre": 0,
        "type": 1,
        "marque": 2,
        "modele": 3,
        "numero_serie": 4,
        "etat": 5,
        "localisation": 6,
        "utilisateur": 7,
        "date_ajout": 8,
        "commentaire": 9,
    }

    desired_items: List[Dict[str, str]] = []
    seen_codes: set[str] = set()
    duplicates = 0

    for row in rows[start_idx:]:
        item = _extract_row(row, idx_map)
        if not item["code_barre"]:
            continue
        if item["code_barre"] in seen_codes:
            duplicates += 1
            continue
        seen_codes.add(item["code_barre"])
        item["date_ajout"] = _parse_date_like_excel(item["date_ajout"])
        desired_items.append(item)

    desired_by_code: Dict[str, Dict[str, str]] = {i["code_barre"]: i for i in desired_items}
    desired_codes = set(desired_by_code.keys())

    wb = None
    updated = 0
    inserted = 0
    try:
        wb = _load_workbook_keep_vba()
        ws = wb[SHEET_NAME]

        max_row = ws.max_row or 0
        code_to_row: Dict[str, int] = {}
        for r in range(HEADER_ROW + 1, max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v:
                c = _norm_code(v)
                if c:
                    code_to_row[c] = r

        to_delete: List[int] = []
        for c, r in code_to_row.items():
            if c not in desired_codes:
                to_delete.append(r)
        for r in sorted(to_delete, reverse=True):
            ws.delete_rows(r, 1)

        ws2 = wb[SHEET_NAME]
        max_row2 = ws2.max_row or 0
        code_to_row2: Dict[str, int] = {}
        for r in range(HEADER_ROW + 1, max_row2 + 1):
            v = ws2.cell(row=r, column=1).value
            if v:
                c = _norm_code(v)
                if c:
                    code_to_row2[c] = r

        for code, item in desired_by_code.items():
            if code in code_to_row2:
                r = code_to_row2[code]
                ws2.cell(row=r, column=1).value = item["code_barre"]
                ws2.cell(row=r, column=2).value = item["type"]
                ws2.cell(row=r, column=3).value = item["marque"]
                ws2.cell(row=r, column=4).value = item["modele"]
                ws2.cell(row=r, column=5).value = item["numero_serie"]
                ws2.cell(row=r, column=6).value = item["etat"]
                ws2.cell(row=r, column=7).value = item["localisation"]
                ws2.cell(row=r, column=8).value = item["utilisateur"]
                ws2.cell(row=r, column=9).value = item["date_ajout"]
                ws2.cell(row=r, column=10).value = item["commentaire"]
                updated += 1
            else:
                ws2.append(
                    [
                        item["code_barre"],
                        item["type"],
                        item["marque"],
                        item["modele"],
                        item["numero_serie"],
                        item["etat"],
                        item["localisation"],
                        item["utilisateur"],
                        item["date_ajout"],
                        item["commentaire"],
                    ]
                )
                inserted += 1

        wb.save(EXCEL_PATH)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    print(
        "OK - Sync Excel termine.\n"
        f"CSV: {csv_path.name}\n"
        f"Maj: {updated} lignes, Ajouts: {inserted} lignes, Codes dupliques ignorés: {duplicates}"
    )


if __name__ == "__main__":
    main()

