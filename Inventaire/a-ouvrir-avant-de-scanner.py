# -*- coding: utf-8 -*-
"""
Ajout rapide matériel avec scan code-barres
- Anti-doublons automatique
- Date automatique
"""

from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import sys

BASE_DIR = Path(__file__).resolve().parent

def charger_codes_existants(ws):
    """Récupère tous les codes-barres déjà dans l'inventaire"""
    codes = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            codes.add(str(row[0]).strip().upper())
    return codes

def main():
    FILE_PATH = BASE_DIR / "inventaire_parc_informatique.xlsx"
    PASSWORD_FILE = BASE_DIR / "mdp_excel.txt"
    PASSWORD = None
    if PASSWORD_FILE.exists():
        try:
            lines = PASSWORD_FILE.read_text(encoding="utf-8").splitlines()
            pwd = next((l.strip() for l in reversed(lines) if l.strip()), "")
            PASSWORD = pwd or None
        except Exception:
            PASSWORD = None
    
    print("\n" + "="*60)
    print("GESTIONNAIRE D'INVENTAIRE")
    print("="*60 + "\n")
    
    if not FILE_PATH.exists():
        print(f"ERREUR : Fichier '{FILE_PATH.name}' introuvable !")
        input("\nAppuyez sur Entree pour quitter...")
        sys.exit(1)
    
    try:
        try:
            wb = load_workbook(FILE_PATH, keep_vba=False)
        except Exception:
            if PASSWORD:
                wb = load_workbook(FILE_PATH, keep_vba=False, password=PASSWORD)
            else:
                raise

        if "Inventaire" not in wb.sheetnames:
            raise KeyError("Feuille 'Inventaire' introuvable dans le classeur.")

        ws = wb["Inventaire"]
    except PermissionError:
        print("ERREUR : Fichier deja ouvert ! Fermez-le et relancez.")
        input("\nAppuyez sur Entree pour quitter...")
        sys.exit(1)
    except Exception as e:
        print(f"ERREUR : {e}")
        input("\nAppuyez sur Entree pour quitter...")
        sys.exit(1)
    
    codes_existants = charger_codes_existants(ws)
    print(f"{len(codes_existants)} codes-barres charges\n")
    
    while True:
        print("="*60)
        
        print("SCAN DU CODE-BARRES")
        code_barre = input(">> Scannez le code-barres (ou 'Q' pour quitter) : ").strip().upper()
        
        if code_barre == 'Q':
            print("\nFermeture du programme...")
            break
        
        if code_barre in codes_existants:
            print("\n" + "="*60)
            print("ATTENTION : CODE-BARRES DEJA EXISTANT !")
            print(f"Code scanne : {code_barre}")
            print("="*60 + "\n")
            continuer = input("Voulez-vous scanner un autre code ? (O/n) : ").strip().upper()
            if continuer in ['N', 'NON']:
                break
            continue
        
        if not code_barre:
            print("\nERREUR : Code-barres vide !")
            continuer = input("Voulez-vous reessayer ? (O/n) : ").strip().upper()
            if continuer in ['N', 'NON']:
                break
            continue
        
        print(f"Code accepte : {code_barre}\n")
        print("="*60)
        
        print("INFORMATIONS DU MATERIEL\n")
        
        type_materiel = input(">> Type (PC/Ecran/Imprimante/Autre) : ").strip()
        marque = input(">> Marque : ").strip()
        modele = input(">> Modele : ").strip()
        numero_serie = input(">> N° serie (optionnel) : ").strip()
        etat = input(">> Etat (Neuf/Bon/Moyen/A reparer) : ").strip()
        localisation = input(">> Localisation : ").strip()
        utilisateur = input(">> Utilisateur (optionnel) : ").strip()
        commentaire = input(">> Commentaire (optionnel) : ").strip()
        
        date_ajout = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        print("\n" + "="*60)
        print("RESUME")
        print("="*60)
        print(f"Code-barres  : {code_barre}")
        print(f"Type         : {type_materiel}")
        print(f"Marque       : {marque}")
        print(f"Modele       : {modele}")
        print(f"N° serie     : {numero_serie or '-'}")
        print(f"Etat         : {etat}")
        print(f"Localisation : {localisation}")
        print(f"Utilisateur  : {utilisateur or '-'}")
        print(f"Date         : {date_ajout} (AUTO)")
        print(f"Commentaire  : {commentaire or '-'}")
        print("="*60 + "\n")
        
        confirmation = input("Valider l'ajout ? (O/n) : ").strip().upper()
        
        if confirmation in ['O', 'OUI', '']:
            try:
                ws.append([
                    code_barre,
                    type_materiel,
                    marque,
                    modele,
                    numero_serie,
                    etat,
                    localisation,
                    utilisateur,
                    date_ajout,
                    commentaire
                ])
                
                wb.save(FILE_PATH)
                
                # Ajouter le code aux codes existants
                codes_existants.add(code_barre)
                
                print("\n" + "="*60)
                print("MATERIEL AJOUTE AVEC SUCCES !")
                print(f"Total : {len(codes_existants)} equipements")
                print("="*60 + "\n")
                
            except PermissionError:
                print("\nERREUR : IMPOSSIBLE DE SAUVEGARDER !")
                print("Le fichier est ouvert ailleurs. Fermez-le et reessayez.\n")
                input("\nAppuyez sur Entree pour continuer...")
                continue
            except Exception as e:
                print(f"\nERREUR lors de la sauvegarde : {e}\n")
                input("\nAppuyez sur Entree pour continuer...")
                continue
        else:
            print("\nAjout annule\n")
        
        continuer = input("Ajouter un autre materiel ? (O/n) : ").strip().upper()
        if continuer in ['N', 'NON']:
            print("\nFermeture du programme...")
            break
        print("\n")
    
    print("\nAu revoir !\n")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation annulee\n")
        input("Appuyez sur Entree pour quitter...")
        sys.exit(0)
    except Exception as e:
        print(f"\nERREUR CRITIQUE : {e}\n")
        input("Appuyez sur Entree pour quitter...")
        sys.exit(1)